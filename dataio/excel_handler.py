# -*- coding: utf-8 -*-
"""
Excel文件处理器

支持导入导出拉表文件
"""

import pandas as pd
from typing import Optional, List
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.models import Skill, Character, Team, Environment


class ExcelHandler:
    """Excel文件处理器"""
    
    def __init__(self):
        self.current_file: Optional[str] = None
    
    def load(self, file_path: str) -> Team:
        """
        从Excel加载队伍
        
        支持拉表模板格式
        """
        self.current_file = file_path
        
        # 使用data_only=True读取计算后的值
        df = pd.read_excel(file_path, sheet_name=0, header=None)
        
        team = Team(name=Path(file_path).stem)
        
        # 查找所有角色分块
        char_starts = self._find_character_blocks(df)
        
        # 解析每个角色
        for start_row in char_starts:
            char = self._parse_character(df, start_row)
            if char and char.skills:
                team.characters.append(char)
        
        return team
    
    def _find_character_blocks(self, df: pd.DataFrame) -> List[int]:
        """查找所有角色分块的起始行"""
        blocks = []
        for i in range(len(df)):
            row = df.iloc[i]
            for j, val in enumerate(row):
                if pd.notna(val) and str(val).strip() == "角色主类型":
                    blocks.append(i)
                    break
        return blocks
    
    def _parse_character(self, df: pd.DataFrame, start_row: int) -> Optional[Character]:
        """解析单个角色"""
        try:
            # 获取角色名（第2行，第2列）
            char_name = ""
            if start_row + 1 < len(df):
                name_val = df.iloc[start_row + 1, 1]
                if pd.notna(name_val):
                    char_name = str(name_val).strip()
            
            if not char_name:
                return None
            
            char = Character(name=char_name)
            
            # 查找技能数据起始行
            skill_start = self._find_skill_start(df, start_row)
            if not skill_start:
                return char
            
            # 解析技能数据
            for i in range(skill_start, min(skill_start + 100, len(df))):
                row = df.iloc[i]
                
                # 检查是否是空行或新角色
                if len(row) <= 14:
                    continue
                
                # 检查第15列（次数）
                count_val = row.iloc[14] if len(row) > 14 else None
                if pd.isna(count_val):
                    # 检查是否是新角色标记
                    if len(row) > 1 and pd.notna(row.iloc[1]):
                        val = str(row.iloc[1]).strip()
                        if val == "角色主类型":
                            break
                    continue
                
                try:
                    skill = self._parse_skill_row(row)
                    if skill and skill.multiplier_input > 0:
                        char.skills.append(skill)
                except Exception as e:
                    continue
            
            return char
            
        except Exception as e:
            print(f"解析角色失败: {e}")
            return None
    
    def _find_skill_start(self, df: pd.DataFrame, start_row: int) -> Optional[int]:
        """查找技能数据起始行"""
        for i in range(start_row + 1, min(start_row + 15, len(df))):
            row = df.iloc[i]
            for j, val in enumerate(row):
                if pd.notna(val) and str(val).strip() == "次数":
                    return i + 1
        return None
    
    def _parse_skill_row(self, row: pd.Series) -> Optional[Skill]:
        """解析技能行数据"""
        skill = Skill()
        
        # 次数 (列15, 索引14)
        if len(row) > 14 and pd.notna(row.iloc[14]):
            skill.count = int(float(row.iloc[14]))
        
        # 动作名称 (列16, 索引15)
        if len(row) > 15 and pd.notna(row.iloc[15]):
            skill.name = str(row.iloc[15]).strip()
        
        # 倍率 (列17, 索引16)
        if len(row) > 16 and pd.notna(row.iloc[16]):
            val = row.iloc[16]
            if isinstance(val, str):
                # 处理带%的字符串
                val = val.replace('%', '').replace('％', '')
                skill.multiplier_input = float(val)
            else:
                # pandas将百分比读取为浮点数（如5.0516表示505.16%）
                val = float(val)
                if val < 10:
                    val = val * 100
                skill.multiplier_input = val
        
        # 类型 (列18, 索引17)
        if len(row) > 17 and pd.notna(row.iloc[17]):
            skill.skill_type = str(row.iloc[17]).strip()
        
        # 面板攻击 (列19, 索引18)
        if len(row) > 18 and pd.notna(row.iloc[18]):
            skill.panel_atk_input = float(row.iloc[18])
        
        # 攻击区 (列20, 索引19) - 处理百分比
        if len(row) > 19 and pd.notna(row.iloc[19]):
            val = str(row.iloc[19]).strip()
            if '%' in val:
                val = val.replace('%', '')
                skill.atk_zone = 1 + float(val) / 100
            else:
                skill.atk_zone = float(val)
        
        # 加成区 (列21, 索引20)
        if len(row) > 20 and pd.notna(row.iloc[20]):
            val = str(row.iloc[20]).strip()
            if '%' in val:
                val = val.replace('%', '')
                skill.bonus_zone = 1 + float(val) / 100
            else:
                skill.bonus_zone = float(val)
        
        # 双爆区 (列22, 索引21)
        if len(row) > 21 and pd.notna(row.iloc[21]):
            skill.crit_zone = float(row.iloc[21])
        
        # 加深区 (列23, 索引22)
        if len(row) > 22 and pd.notna(row.iloc[22]):
            skill.amplify_zone = float(row.iloc[22])
        
        # 防御区 (列24, 索引23)
        if len(row) > 23 and pd.notna(row.iloc[23]):
            skill.defense_zone = float(row.iloc[23])
        
        # 抗性区 (列25, 索引24)
        if len(row) > 24 and pd.notna(row.iloc[24]):
            skill.resistance_zone = float(row.iloc[24])
        
        # 倍率提升 (列26, 索引25)
        if len(row) > 25 and pd.notna(row.iloc[25]):
            val = str(row.iloc[25]).strip().replace('%', '')
            skill.multiplier_boost = float(val) / 100
        
        # 易伤区 (列27, 索引26)
        if len(row) > 26 and pd.notna(row.iloc[26]):
            val = str(row.iloc[26]).strip().replace('%', '')
            skill.vulnerable_zone = float(val) / 100
        
        # 独立乘区 (列28, 索引27)
        if len(row) > 27 and pd.notna(row.iloc[27]):
            val = str(row.iloc[27]).strip().replace('%', '')
            skill.independent_zone = 1 + float(val) / 100
        
        # 伤害 (列29, 索引28)
        if len(row) > 28 and pd.notna(row.iloc[28]):
            skill.damage = float(row.iloc[28])
        
        # 余响 (列30, 索引29)
        if len(row) > 29 and pd.notna(row.iloc[29]):
            skill.echo = int(float(row.iloc[29]))
        
        # 溢出 (列31, 索引30)
        if len(row) > 30 and pd.notna(row.iloc[30]):
            skill.overflow = int(float(row.iloc[30]))
        
        return skill
    
    def save(self, team: Team, file_path: str):
        """
        保存队伍到Excel
        
        输出标准拉表格式
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = team.name if team.name else "队伍"
        
        # 设置列宽
        col_widths = [3, 12, 8, 12, 10, 12, 10, 12, 10, 12, 10, 12, 10, 3,
                      6, 15, 10, 6, 12, 10, 10, 10, 10, 10, 10, 10, 10, 10, 12, 6, 6]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 样式定义
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        skill_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        current_row = 1
        
        for char in team.characters:
            # 角色标记行
            ws.cell(row=current_row, column=2, value="角色主类型")
            ws.cell(row=current_row, column=3, value="e")
            ws.cell(row=current_row, column=4, value="3C属伤数")
            ws.cell(row=current_row, column=5, value=2)
            current_row += 1
            
            # 角色名行 + 技能标题
            ws.cell(row=current_row, column=2, value=char.name)
            headers = ["次数", "动作", "倍率", "类型", "面板攻击", "攻击区", "加成区", 
                      "双爆区", "加深区", "防御区", "抗性区", "倍率提升", "易伤区", 
                      "独立乘区", "伤害", "余响", "溢出"]
            for i, header in enumerate(headers, 15):
                cell = ws.cell(row=current_row, column=i, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            current_row += 1
            
            # 属性标签行
            ws.cell(row=current_row, column=2, value="攻击")
            ws.cell(row=current_row, column=4, value="加成区")
            ws.cell(row=current_row, column=6, value="暴击")
            ws.cell(row=current_row, column=8, value="暴击伤害")
            ws.cell(row=current_row, column=10, value="加深区")
            ws.cell(row=current_row, column=12, value="副词条")
            current_row += 1
            
            # 技能数据行
            for skill in char.skills:
                # 计算显示值
                atk_zone_display = f"{(skill.atk_zone - 1) * 100:.2f}%" if skill.atk_zone != 1 else ""
                bonus_zone_display = f"{(skill.bonus_zone - 1) * 100:.2f}%" if skill.bonus_zone != 1 else ""
                
                ws.cell(row=current_row, column=15, value=skill.count)
                ws.cell(row=current_row, column=16, value=skill.name)
                ws.cell(row=current_row, column=17, value=f"{skill.multiplier_input:.2f}%")
                ws.cell(row=current_row, column=18, value=skill.skill_type)
                ws.cell(row=current_row, column=19, value=int(skill.panel_atk_input) if skill.panel_atk_input else "")
                ws.cell(row=current_row, column=20, value=atk_zone_display)
                ws.cell(row=current_row, column=21, value=bonus_zone_display)
                ws.cell(row=current_row, column=22, value=f"{skill.crit_zone:.4f}" if skill.crit_zone else "")
                ws.cell(row=current_row, column=23, value=skill.amplify_zone)
                ws.cell(row=current_row, column=24, value=f"{skill.defense_zone:.4f}" if skill.defense_zone else "")
                ws.cell(row=current_row, column=25, value=skill.resistance_zone)
                ws.cell(row=current_row, column=26, value=f"{skill.multiplier_boost * 100:.0f}%")
                ws.cell(row=current_row, column=27, value=f"{skill.vulnerable_zone * 100:.0f}%")
                ws.cell(row=current_row, column=28, value=f"{(skill.independent_zone - 1) * 100:.0f}%")
                ws.cell(row=current_row, column=29, value=int(skill.damage) if skill.damage else "")
                ws.cell(row=current_row, column=30, value=skill.echo)
                ws.cell(row=current_row, column=31, value=skill.overflow)
                
                # 应用样式
                for col in range(15, 32):
                    ws.cell(row=current_row, column=col).fill = skill_fill
                
                current_row += 1
            
            # 空行分隔
            current_row += 1
        
        wb.save(file_path)
        print(f"已保存到: {file_path}")
