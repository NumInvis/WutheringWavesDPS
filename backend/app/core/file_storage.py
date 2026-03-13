"""
File storage service.
Supports local storage (and leaves MinIO/S3 for future).
Enhanced with Excel preservation capabilities.
"""
import os
import uuid
import hashlib
from typing import Optional, Tuple, Dict, Any
from datetime import datetime

from fastapi import UploadFile, HTTPException, status

from app.core.config import get_settings

settings = get_settings()


class FileStorageService:
    """File storage service with Excel preservation support."""

    def __init__(self):
        self.storage_type = settings.storage_type or "local"
        self.upload_dir = settings.upload_dir or "./uploads"
        self.excel_storage_dir = os.path.join(os.path.dirname(self.upload_dir), "excel_storage")

        if self.storage_type == "local":
            os.makedirs(self.upload_dir, exist_ok=True)
            os.makedirs(self.excel_storage_dir, exist_ok=True)

    def _generate_filename(self, original_filename: str) -> str:
        """Generate a unique filename."""
        ext = os.path.splitext(original_filename)[1]
        return f"{uuid.uuid4()}{ext}"

    def _generate_file_id(self) -> str:
        """Generate a unique file ID."""
        return str(uuid.uuid4())

    async def save_file(self, file: UploadFile, max_size: Optional[int] = None) -> tuple[str, int]:
        """
        Save file (binary preservation).

        Returns:
            (file_url, file_size)
        """
        content = await file.read()
        file_size = len(content)

        if max_size and file_size > max_size:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=f"文件大小超过限制（最大 {max_size / 1024 / 1024:.1f}MB）"
            )

        filename = self._generate_filename(file.filename or "unknown")

        if self.storage_type == "local":
            file_path = os.path.join(self.upload_dir, filename)
            with open(file_path, "wb") as f:
                f.write(content)

            file_url = f"/uploads/{filename}"
            return file_url, file_size

        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail=f"存储类型 {self.storage_type} 尚未实现"
        )

    async def save_excel_with_preservation(
        self, 
        file: UploadFile, 
        max_size: Optional[int] = None
    ) -> Tuple[str, str, int, Dict[str, Any]]:
        """
        Save Excel file with full preservation (data, formulas, formats).
        
        Returns:
            (file_id, file_url, file_size, metadata)
        """
        content = await file.read()
        file_size = len(content)

        if max_size and file_size > max_size:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=f"文件大小超过限制（最大 {max_size / 1024 / 1024:.1f}MB）"
            )

        ext = os.path.splitext(file.filename or "unknown.xlsx")[1].lower()
        if ext not in [".xlsx", ".xlsm", ".xls"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"不支持的Excel格式: {ext}，仅支持 .xlsx, .xlsm, .xls"
            )

        file_id = self._generate_file_id()
        file_hash = hashlib.sha256(content).hexdigest()

        if self.storage_type == "local":
            original_path = os.path.join(self.excel_storage_dir, f"{file_id}_original{ext}")
            with open(original_path, "wb") as f:
                f.write(content)

            metadata = self._extract_excel_metadata(content, file.filename, file_hash, file_size)
            
            metadata_path = os.path.join(self.excel_storage_dir, f"{file_id}.meta.json")
            import json
            with open(metadata_path, "w", encoding="utf-8") as f:
                json.dump(metadata, f, ensure_ascii=False, indent=2)

            access_path = os.path.join(self.upload_dir, f"{file_id}{ext}")
            with open(access_path, "wb") as f:
                f.write(content)

            file_url = f"/uploads/{file_id}{ext}"
            return file_id, file_url, file_size, metadata

        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail=f"存储类型 {self.storage_type} 尚未实现"
        )

    def _extract_excel_metadata(
        self, 
        content: bytes, 
        filename: str, 
        file_hash: str,
        file_size: int
    ) -> Dict[str, Any]:
        """Extract Excel file metadata."""
        import io
        from openpyxl import load_workbook
        
        metadata = {
            "filename": filename,
            "file_size": file_size,
            "file_hash": file_hash,
            "sheet_count": 0,
            "sheet_names": [],
            "created_at": datetime.now().isoformat(),
            "modified_at": None,
            "author": None,
            "last_modified_by": None,
            "has_formulas": False,
            "formula_count": 0,
            "cell_count": 0,
            "merged_cell_ranges": 0
        }
        
        try:
            with io.BytesIO(content) as buffer:
                wb = load_workbook(buffer, read_only=True, data_only=False)
                
                metadata["sheet_names"] = wb.sheetnames
                metadata["sheet_count"] = len(wb.sheetnames)
                
                props = wb.properties
                if props.created:
                    metadata["created_at"] = props.created.isoformat()
                if props.modified:
                    metadata["modified_at"] = props.modified.isoformat()
                if props.creator:
                    metadata["author"] = props.creator
                if props.lastModifiedBy:
                    metadata["last_modified_by"] = props.lastModifiedBy
                
                formula_count = 0
                cell_count = 0
                merged_count = 0
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    merged_count += len(ws.merged_cells.ranges)
                    
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                cell_count += 1
                                if cell.data_type == 'f':
                                    formula_count += 1
                
                metadata["has_formulas"] = formula_count > 0
                metadata["formula_count"] = formula_count
                metadata["cell_count"] = cell_count
                metadata["merged_cell_ranges"] = merged_count
                
                wb.close()
        except Exception as e:
            metadata["parse_error"] = str(e)
        
        return metadata

    def get_excel_original(self, file_id: str) -> Tuple[bytes, Dict[str, Any]]:
        """
        Get original Excel file content (binary, unmodified).
        
        Returns:
            (content, metadata)
        """
        import json
        
        metadata_path = os.path.join(self.excel_storage_dir, f"{file_id}.meta.json")
        if not os.path.exists(metadata_path):
            raise FileNotFoundError(f"找不到文件元数据: {file_id}")
        
        with open(metadata_path, "r", encoding="utf-8") as f:
            metadata = json.load(f)
        
        for ext in [".xlsx", ".xlsm", ".xls"]:
            file_path = os.path.join(self.excel_storage_dir, f"{file_id}_original{ext}")
            if os.path.exists(file_path):
                with open(file_path, "rb") as f:
                    return f.read(), metadata
        
        raise FileNotFoundError(f"找不到原始文件: {file_id}")

    def delete_file(self, file_url: str) -> bool:
        """Delete file."""
        if self.storage_type == "local":
            filename = os.path.basename(file_url)
            file_path = os.path.join(self.upload_dir, filename)
            if os.path.exists(file_path):
                os.remove(file_path)
                return True
            return False

        return False

    def delete_excel_file(self, file_id: str) -> bool:
        """Delete Excel file and its metadata."""
        deleted = False
        
        for ext in ["_original.xlsx", "_original.xlsm", "_original.xls", ".meta.json"]:
            file_path = os.path.join(self.excel_storage_dir, f"{file_id}{ext}")
            if os.path.exists(file_path):
                os.remove(file_path)
                deleted = True
        
        for ext in [".xlsx", ".xlsm", ".xls"]:
            access_path = os.path.join(self.upload_dir, f"{file_id}{ext}")
            if os.path.exists(access_path):
                os.remove(access_path)
                deleted = True
        
        return deleted


file_storage = FileStorageService()
