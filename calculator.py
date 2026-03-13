# -*- coding: utf-8 -*-
"""
鸣潮DPS计算器 v1.0 - 核心计算器
"""

import pandas as pd
from dataclasses import dataclass, field
from typing import List, Dict, Optional
from pathlib import Path


@dataclass
class Environment:
    """环境设置"""
    name: str = "默认环境"
    enemy_level: int = 100  # 怪物等级
    resistance_zone: float = 0.8  # 抗性区（直接设置）
    
    def get_defense_zone(self, def_shred: float = 0, def_ignore: float = 0) -> float:
        """计算防御区
        
        公式: 1520 / (1520 + (792 + 8*怪物等级) * (1-减防%) * (1-无视防御%))
        """
        numerator = 1520
        denominator = 1520 + (792 + 8 * self.enemy_level) * (1 - def_shred) * (1 - def_ignore)
        return numerator / denominator


@dataclass
class Skill:
    """技能/动作数据"""
    # 基础信息
    count: int = 1  # 次数
    name: str = ""  # 动作名称
    multiplier: float = 0  # 倍率(%)
    skill_type: str = "a"  # 类型 (a/e/q等)
    
    # 伤害乘区
    panel_atk: float = 0  # 面板攻击
    atk_zone: float = 1.0  # 攻击区
    bonus_zone: float = 1.0  # 加成区
    crit_zone: float = 1.0  # 双爆区
    amplify_zone: float = 0  # 加深区
    defense_zone: float = 0.5  # 防御区
    resistance_zone: float = 0.8  # 抗性区
    
    # 特殊乘区
    multiplier_boost: float = 0  # 倍率提升
    vulnerable_zone: float = 0  # 易伤区
    independent_zone: float = 1.0  # 独立乘区
    
    # 计算结果
    damage: float = 0  # 单次伤害
    total_damage: float = 0  # 总伤害 (伤害 * 次数)
    
    # 其他
    echo: int = 0  # 余响
    overflow: int = 0  # 溢出
    
    def calculate(self) -> float:
        """计算伤害"""
        self.damage = (
            self.panel_atk * 
            (self.multiplier / 100) * 
            self.atk_zone * 
            self.bonus_zone * 
            self.crit_zone * 
            (1 + self.amplify_zone) * 
            self.defense_zone * 
            self.resistance_zone * 
            (1 + self.multiplier_boost) * 
            (1 + self.vulnerable_zone) * 
            self.independent_zone
        )
        self.total_damage = self.damage * self.count
        return self.damage


@dataclass
class Character:
    """角色"""
    name: str = ""
    skills: List[Skill] = field(default_factory=list)
    
    def get_total_damage(self) -> float:
        """获取角色总伤害"""
        return sum(s.total_damage for s in self.skills)


@dataclass
class Team:
    """队伍"""
    name: str = ""
    characters: List[Character] = field(default_factory=list)
    environment: Environment = field(default_factory=Environment)
    time_seconds: float = 25.0  # 时间（秒）
    
    def get_team_damage(self) -> float:
        """获取队伍总伤害"""
        return sum(c.get_total_damage() for c in self.characters)
    
    def get_team_dps(self) -> float:
        """获取队伍DPS"""
        return self.get_team_damage() / self.time_seconds if self.time_seconds > 0 else 0


class WuWaCalculator:
    """鸣潮计算器"""
    
    def __init__(self):
        self.teams: Dict[str, Team] = {}
        self.current_team: Optional[Team] = None
    
    def load_from_excel(self, file_path: str) -> Team:
        """从Excel加载队伍 - 使用data_only=True读取计算值"""
        # 使用data_only=True读取计算后的值
        df = pd.read_excel(file_path, sheet_name=0, header=None)
        
        team = Team(name=Path(file_path).stem)
        
        # 查找所有角色分块（通过"角色主类型"标记）
        char_starts = []
        for i in range(len(df)):
            row = df.iloc[i]
            for j, val in enumerate(row):
                if pd.notna(val) and str(val).strip() == "角色主类型":
                    char_starts.append(i)
                    break
        
        # 解析每个角色
        for idx, start_row in enumerate(char_starts):
            char = self._parse_character(df, start_row)
            if char and char.skills:
                team.characters.append(char)
        
        self.teams[team.name] = team
        self.current_team = team
        
        return team
    
    def _parse_character(self, df: pd.DataFrame, start_row: int) -> Optional[Character]:
        """解析角色"""
        try:
            # 获取角色名（第2行，第2列）
            char_name = ""
            if start_row + 1 < len(df):
                name_val = df.iloc[start_row + 1, 1]
                if pd.notna(name_val):
                    char_name = str(name_val).strip()
            
            if not char_name:
                return None
            
            char = Character(name=char_name)
            
            # 查找"次数"标题行
            header_row = None
            for i in range(start_row + 1, min(start_row + 5, len(df))):
                row = df.iloc[i]
                for j, val in enumerate(row):
                    if pd.notna(val) and str(val).strip() == "次数":
                        header_row = i
                        break
                if header_row:
                    break
            
            if not header_row:
                return char
            
            # 解析技能数据（从标题行+1开始）
            for i in range(header_row + 1, min(header_row + 50, len(df))):
                row = df.iloc[i]
                
                # 检查是否是空行或新角色开始
                if len(row) <= 14:
                    continue
                
                # 检查第15列（次数）是否有有效数据
                count_val = row.iloc[14] if len(row) > 14 else None
                if pd.isna(count_val):
                    # 检查是否是新角色标记
                    if len(row) > 1 and pd.notna(row.iloc[1]):
                        val = str(row.iloc[1]).strip()
                        if val == "角色主类型":
                            break
                    continue
                
                try:
                    skill = Skill()
                    
                    # 次数 (列15)
                    skill.count = int(float(count_val))
                    
                    # 动作名称 (列16)
                    if len(row) > 15 and pd.notna(row.iloc[15]):
                        skill.name = str(row.iloc[15]).strip()
                    
                    # 倍率 (列17) - 处理带%的情况
                    if len(row) > 16 and pd.notna(row.iloc[16]):
                        val = str(row.iloc[16]).strip()
                        if '%' in val:
                            val = val.replace('%', '')
                        skill.multiplier = float(val)
                    
                    # 类型 (列18)
                    if len(row) > 17 and pd.notna(row.iloc[17]):
                        skill.skill_type = str(row.iloc[17]).strip()
                    
                    # 面板攻击 (列19)
                    if len(row) > 18 and pd.notna(row.iloc[18]):
                        skill.panel_atk = float(row.iloc[18])
                    
                    # 攻击区 (列20)
                    if len(row) > 19 and pd.notna(row.iloc[19]):
                        val = str(row.iloc[19]).strip()
                        if '%' in val:
                            val = val.replace('%', '')
                            skill.atk_zone = 1 + float(val) / 100
                        else:
                            skill.atk_zone = float(val)
                    
                    # 加成区 (列21)
                    if len(row) > 20 and pd.notna(row.iloc[20]):
                        val = str(row.iloc[20]).strip()
                        if '%' in val:
                            val = val.replace('%', '')
                            skill.bonus_zone = 1 + float(val) / 100
                        else:
                            skill.bonus_zone = float(val)
                    
                    # 双爆区 (列22)
                    if len(row) > 21 and pd.notna(row.iloc[21]):
                        skill.crit_zone = float(row.iloc[21])
                    
                    # 加深区 (列23)
                    if len(row) > 22 and pd.notna(row.iloc[22]):
                        skill.amplify_zone = float(row.iloc[22])
                    
                    # 防御区 (列24)
                    if len(row) > 23 and pd.notna(row.iloc[23]):
                        skill.defense_zone = float(row.iloc[23])
                    
                    # 抗性区 (列25)
                    if len(row) > 24 and pd.notna(row.iloc[24]):
                        skill.resistance_zone = float(row.iloc[24])
                    
                    # 倍率提升 (列26)
                    if len(row) > 25 and pd.notna(row.iloc[25]):
                        val = str(row.iloc[25]).strip()
                        if '%' in val:
                            val = val.replace('%', '')
                        skill.multiplier_boost = float(val) / 100
                    
                    # 易伤区 (列27)
                    if len(row) > 26 and pd.notna(row.iloc[26]):
                        val = str(row.iloc[26]).strip()
                        if '%' in val:
                            val = val.replace('%', '')
                        skill.vulnerable_zone = float(val) / 100
                    
                    # 独立乘区 (列28)
                    if len(row) > 27 and pd.notna(row.iloc[27]):
                        val = str(row.iloc[27]).strip()
                        if '%' in val:
                            val = val.replace('%', '')
                        skill.independent_zone = 1 + float(val) / 100
                    
                    # 伤害 (列29)
                    if len(row) > 28 and pd.notna(row.iloc[28]):
                        skill.damage = float(row.iloc[28])
                    
                    # 余响 (列30)
                    if len(row) > 29 and pd.notna(row.iloc[29]):
                        skill.echo = int(float(row.iloc[29]))
                    
                    # 溢出 (列31)
                    if len(row) > 30 and pd.notna(row.iloc[30]):
                        skill.overflow = int(float(row.iloc[30]))
                    
                    # 计算总伤害
                    skill.total_damage = skill.damage * skill.count
                    
                    char.skills.append(skill)
                    
                except Exception as e:
                    continue
            
            return char
            
        except Exception as e:
            print(f"解析角色失败: {e}")
            return None
    
    def save_to_excel(self, team: Team, output_path: str):
        """保存队伍到Excel - 表格形式"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = team.name if team.name else "队伍"
        
        # 设置列宽
        col_widths = [3, 12, 8, 12, 10, 12, 10, 12, 10, 12, 10, 12, 10, 3,
                      6, 15, 10, 6, 12, 10, 10, 10, 10, 10, 10, 10, 10, 10, 12, 6, 6]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        current_row = 1
        
        for char in team.characters:
            # 角色标记行
            ws.cell(row=current_row, column=2, value="角色主类型")
            ws.cell(row=current_row, column=3, value="e")
            ws.cell(row=current_row, column=4, value="3C属伤数")
            ws.cell(row=current_row, column=5, value=2)
            current_row += 1
            
            # 角色名行 + 技能标题
            ws.cell(row=current_row, column=2, value=char.name)
            ws.cell(row=current_row, column=15, value="次数")
            ws.cell(row=current_row, column=16, value="动作")
            ws.cell(row=current_row, column=17, value="倍率")
            ws.cell(row=current_row, column=18, value="类型")
            ws.cell(row=current_row, column=19, value="面板攻击")
            ws.cell(row=current_row, column=20, value="攻击区")
            ws.cell(row=current_row, column=21, value="加成区")
            ws.cell(row=current_row, column=22, value="双爆区")
            ws.cell(row=current_row, column=23, value="加深区")
            ws.cell(row=current_row, column=24, value="防御区")
            ws.cell(row=current_row, column=25, value="抗性区")
            ws.cell(row=current_row, column=26, value="倍率提升")
            ws.cell(row=current_row, column=27, value="易伤区")
            ws.cell(row=current_row, column=28, value="独立乘区")
            ws.cell(row=current_row, column=29, value="伤害")
            ws.cell(row=current_row, column=30, value="余响")
            ws.cell(row=current_row, column=31, value="溢出")
            current_row += 1
            
            # 属性标签行
            ws.cell(row=current_row, column=2, value="攻击")
            ws.cell(row=current_row, column=4, value="加成区")
            ws.cell(row=current_row, column=6, value="暴击")
            ws.cell(row=current_row, column=8, value="暴击伤害")
            ws.cell(row=current_row, column=10, value="加深区")
            ws.cell(row=current_row, column=12, value="副词条")
            current_row += 1
            
            # 技能数据行
            for skill in char.skills:
                ws.cell(row=current_row, column=15, value=skill.count)
                ws.cell(row=current_row, column=16, value=skill.name)
                ws.cell(row=current_row, column=17, value=f"{skill.multiplier}%")
                ws.cell(row=current_row, column=18, value=skill.skill_type)
                ws.cell(row=current_row, column=19, value=skill.panel_atk)
                ws.cell(row=current_row, column=20, value=f"{(skill.atk_zone - 1) * 100:.2f}%" if skill.atk_zone != 1 else "")
                ws.cell(row=current_row, column=21, value=f"{(skill.bonus_zone - 1) * 100:.2f}%" if skill.bonus_zone != 1 else "")
                ws.cell(row=current_row, column=22, value=skill.crit_zone)
                ws.cell(row=current_row, column=23, value=skill.amplify_zone)
                ws.cell(row=current_row, column=24, value=skill.defense_zone)
                ws.cell(row=current_row, column=25, value=skill.resistance_zone)
                ws.cell(row=current_row, column=26, value=f"{skill.multiplier_boost * 100:.0f}%")
                ws.cell(row=current_row, column=27, value=f"{skill.vulnerable_zone * 100:.0f}%")
                ws.cell(row=current_row, column=28, value=f"{(skill.independent_zone - 1) * 100:.0f}%")
                ws.cell(row=current_row, column=29, value=int(skill.damage))
                ws.cell(row=current_row, column=30, value=skill.echo)
                ws.cell(row=current_row, column=31, value=skill.overflow)
                current_row += 1
            
            # 空行
            current_row += 1
        
        # 保存
        wb.save(output_path)
        print(f"已保存到: {output_path}")
