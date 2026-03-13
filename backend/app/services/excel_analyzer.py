from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook


@dataclass(frozen=True)
class DpsSheetSummary:
    sheet_name: str
    total_damage: Optional[float]
    dps: Optional[float]
    time_seconds: Optional[float]
    total_damage_cell: Optional[str]
    dps_cell: Optional[str]
    time_cell: Optional[str]


_DPS_LABEL_RE = re.compile(r"^\s*dps\s*$", re.IGNORECASE)
_DIV_FORMULA_RE = re.compile(
    r"^\s*=\s*([A-Z]{1,3}\d+)\s*/\s*\$?([A-Z]{1,3})\$?(\d+)\s*$",
    re.IGNORECASE,
)


def _to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except Exception:
        return None


def analyze_dps_summaries(xlsx_path: str | Path) -> list[DpsSheetSummary]:
    """
    从拉表 xlsx 中提取“总伤/DPS/轴长”摘要。
    - 不计算公式，只读取 Excel 缓存值（data_only=True）。
    - 用 data_only=False 定位 dps 单元格与分母（轴长）引用。
    """
    xlsx_path = Path(xlsx_path)
    wb_formula = load_workbook(xlsx_path, data_only=False, read_only=True)
    wb_value = load_workbook(xlsx_path, data_only=True, read_only=True)

    summaries: list[DpsSheetSummary] = []
    try:
        for sheet_name in wb_formula.sheetnames:
            ws_f = wb_formula[sheet_name]
            ws_v = wb_value[sheet_name]

            dps_label_coords: list[tuple[int, int]] = []
            for row in ws_f.iter_rows(values_only=False):
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and _DPS_LABEL_RE.match(v):
                        dps_label_coords.append((cell.row, cell.column))

            found_any = False
            for (r, c) in dps_label_coords:
                dps_cell = ws_f.cell(r + 1, c)
                total_cell = ws_f.cell(r + 1, c - 1) if c > 1 else None

                dps_formula = dps_cell.value if isinstance(dps_cell.value, str) else None
                time_cell_ref: Optional[str] = None
                if dps_formula:
                    m = _DIV_FORMULA_RE.match(dps_formula)
                    if m:
                        col = m.group(2).upper()
                        rown = m.group(3)
                        time_cell_ref = f"{col}{rown}"

                dps_value = _to_float(ws_v[dps_cell.coordinate].value)
                total_value = _to_float(ws_v[total_cell.coordinate].value) if total_cell else None
                time_value = _to_float(ws_v[time_cell_ref].value) if time_cell_ref else None

                summaries.append(
                    DpsSheetSummary(
                        sheet_name=sheet_name,
                        total_damage=total_value,
                        dps=dps_value,
                        time_seconds=time_value,
                        total_damage_cell=total_cell.coordinate if total_cell else None,
                        dps_cell=dps_cell.coordinate,
                        time_cell=time_cell_ref,
                    )
                )
                found_any = True

            if not found_any:
                summaries.append(
                    DpsSheetSummary(
                        sheet_name=sheet_name,
                        total_damage=None,
                        dps=None,
                        time_seconds=None,
                        total_damage_cell=None,
                        dps_cell=None,
                        time_cell=None,
                    )
                )
    finally:
        wb_formula.close()
        wb_value.close()

    return summaries
